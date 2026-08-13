"""
将登录模块测试用例导出为标准 Excel 测试用例文档。
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from utils.test_case_meta import TEST_CASES

OUTPUT = Path(__file__).parent.parent / "docs" / "登录功能测试用例.xlsx"

HEADERS = [
  "用例编号",
  "用例标题",
  "所属模块",
  "功能点",
  "优先级",
  "用例类型",
  "前置条件",
  "测试步骤",
  "预期结果",
  "测试数据",
  "关联自动化用例",
  "备注",
]

COLUMN_WIDTHS = {
  "A": 14,
  "B": 32,
  "C": 12,
  "D": 18,
  "E": 8,
  "F": 10,
  "G": 28,
  "H": 36,
  "I": 36,
  "J": 22,
  "K": 38,
  "L": 14,
}


def build_workbook() -> Workbook:
  wb = Workbook()
  ws = wb.active
  ws.title = "登录功能测试用例"

  header_fill = PatternFill("solid", fgColor="4472C4")
  header_font = Font(bold=True, color="FFFFFF", size=11)
  thin = Side(style="thin", color="B4B4B4")
  border = Border(left=thin, right=thin, top=thin, bottom=thin)
  wrap = Alignment(wrap_text=True, vertical="top")

  for col, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

  for row_idx, case in enumerate(TEST_CASES, start=2):
    values = [
      case["id"],
      case["title"],
      case["module"],
      case["story"],
      case["priority"],
      case["type"],
      case["precondition"],
      case["steps"],
      case["expected"],
      case["data"],
      case["method"],
      case["remark"],
    ]
    for col_idx, value in enumerate(values, start=1):
      cell = ws.cell(row=row_idx, column=col_idx, value=value)
      cell.alignment = wrap
      cell.border = border

  ws.freeze_panes = "A2"
  ws.row_dimensions[1].height = 28
  for row in range(2, len(TEST_CASES) + 2):
    ws.row_dimensions[row].height = 90

  for col_letter, width in COLUMN_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

  return wb


def main():
  OUTPUT.parent.mkdir(parents=True, exist_ok=True)
  wb = build_workbook()
  wb.save(OUTPUT)
  print(f"已生成: {OUTPUT}")


if __name__ == "__main__":
  main()
